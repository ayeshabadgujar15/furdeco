import hashlib
import hmac
import io
import os
import tempfile
import time
from datetime import datetime, timedelta

import extra_streamlit_components as stx
import pandas as pd
import requests
import streamlit as st
from openpyxl.styles import Alignment, PatternFill

from extractor import extract_from_pdf

PRICE_API_URL = "https://www.choicefurnituresuperstore.co.uk/order_furdeco_netTotal.php"

# Some sites quietly reject requests that don't look like they came from a
# real browser — Python's requests library sends a distinctive default
# User-Agent ("python-requests/x.y") that's an easy, common thing for a
# server (or something in front of it, like a WAF) to block, even while the
# exact same order number works fine when a person visits the URL directly.
# A normal browser User-Agent header sidesteps that without changing
# anything about what's actually being requested.
PRICE_API_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
}


@st.cache_data(show_spinner=False, ttl=3600)
def fetch_price(order_no: str):
    """Look up the net total (Price) for one order number from CFS's own
    order_furdeco_netTotal API. Cached by order number for an hour, so
    re-running the script (e.g. every time a filter checkbox is toggled,
    since Streamlit reruns the whole script on any widget interaction)
    doesn't keep re-hitting the API for orders already looked up in this
    session.

    Returns (price, reason): price is None on any failure, and reason is a
    short machine-readable tag saying *why* (empty order number, timeout,
    connection error, non-2xx HTTP status, unparseable JSON, the API's own
    "not found"/error response, or a success response missing/malforming
    the field we need). None of these raise — one bad lookup should never
    take down the whole extraction — but the reason lets the UI show a
    breakdown instead of one opaque "couldn't be looked up" count, which is
    the only way to tell "the order numbers are wrong" apart from "the API
    was unreachable" or "CFS genuinely doesn't have these orders"."""
    if not order_no:
        return None, "empty_order_number"
    try:
        resp = requests.get(
            PRICE_API_URL,
            params={"order_no": order_no},
            headers=PRICE_API_HEADERS,
            timeout=15,
        )
    except requests.Timeout:
        return None, "timeout"
    except requests.RequestException:
        return None, "connection_error"

    if not resp.ok:
        return None, f"http_{resp.status_code}"

    try:
        data = resp.json()
    except ValueError:
        return None, "unparseable_response"

    if data.get("status") != "success":
        return None, "api_reported_not_found"

    try:
        return float(data["fNetTotal"]), None
    except (KeyError, TypeError, ValueError):
        return None, "malformed_success_response"

# --- Palette ---
DARK_GREEN = "#0A3323"
MOSS_GREEN = "#839958"
BEIGE = "#F7F4D5"
ROSY_BROWN = "#D3968C"
MIDNIGHT_GREEN = "#105666"
LIGHT_BLUE = "#ADD8E6"  # flags rows where Price couldn't be looked up

HIGHLIGHT_THRESHOLD = 50  # AMOUNT above this gets the row highlighted


def _blend_with_white(hex_color: str, amount: float) -> str:
    """Lighten a hex color by blending it with white (amount=0 -> white, 1 -> original).
    Used for the Excel row fill, since Excel fills need a solid color rather than
    the translucent overlay used on-screen."""
    hex_color = hex_color.lstrip("#")
    r, g, b = (int(hex_color[i:i + 2], 16) for i in (0, 2, 4))
    r, g, b = (round(255 - (255 - c) * amount) for c in (r, g, b))
    return f"{r:02X}{g:02X}{b:02X}"

st.set_page_config(page_title="Furdeco Invoice Extractor", page_icon="📄", layout="wide")

OUTPUT_COLUMNS = ["DATE", "RECIPIENT", "ORDER #", "POSTCODE", "DETAILS", "AMOUNT", "Price", "Invoice No"]

st.markdown(
    f"""
    <style>
    /* Everything here is layered on top of Streamlit's own light/dark theme
    rather than replacing it — .stApp keeps its normal (theme-controlled)
    background, so switching the app's theme (light/dark, via the Settings
    menu) still works. Only self-contained accent elements (banner, buttons,
    borders) use the fixed palette, each with its own guaranteed-readable
    text color, so they look right regardless of the surrounding theme. */

    .furdeco-banner {{
        background-color: {ROSY_BROWN};
        color: {MIDNIGHT_GREEN};
        padding: 1.4rem 1.8rem;
        border-radius: 10px;
        margin-bottom: 1.4rem;
        border: 1px solid {MIDNIGHT_GREEN}55;
    }}
    .furdeco-banner h1 {{
        color: {MIDNIGHT_GREEN};
        margin: 0;
        font-size: 1.7rem;
    }}
    .furdeco-banner p {{
        color: {MIDNIGHT_GREEN};
        margin: 0.3rem 0 0 0;
        font-size: 0.95rem;
        font-weight: 600;
    }}
    section[data-testid="stFileUploaderDropzone"] {{
        border: 1.5px dashed {MOSS_GREEN} !important;
        border-radius: 8px;
        padding: 0.5rem !important;
    }}
    section[data-testid="stFileUploaderDropzone"] button {{
        padding: 0.15rem 0.7rem !important;
        min-height: 0 !important;
        font-size: 0.8rem !important;
    }}
    section[data-testid="stFileUploaderDropzone"] small {{
        font-size: 0.72rem !important;
    }}
    .stButton > button, .stDownloadButton > button {{
        background-color: {MIDNIGHT_GREEN};
        color: {BEIGE};
        border: 1px solid {MIDNIGHT_GREEN};
        border-radius: 6px;
        font-weight: 600;
    }}
    .stButton > button:hover, .stDownloadButton > button:hover {{
        background-color: {ROSY_BROWN};
        border-color: {ROSY_BROWN};
        color: {DARK_GREEN};
    }}
    div[data-testid="stExpander"] {{
        border: 1px solid {MOSS_GREEN} !important;
        border-radius: 8px;
    }}
    div[data-testid="stExpander"] summary {{
        color: {MOSS_GREEN};
        font-weight: 600;
    }}
    div[data-testid="stMetric"] {{
        border-radius: 8px;
        padding: 0.6rem 0.8rem;
        border: 1px solid {MOSS_GREEN};
        border-top: 3px solid {ROSY_BROWN};
    }}
    [data-testid="stMetricLabel"] {{
        color: {MOSS_GREEN} !important;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="furdeco-banner">
        <h1>Furdeco Invoice Extractor</h1>
        <p>Upload one or more Furdeco invoice PDFs and get a single Excel file with every line item.</p>
    </div>
    """,
    unsafe_allow_html=True,
)


COOKIE_NAME = "furdeco_auth"
COOKIE_EXPIRY_DAYS = 30


def get_cookie_manager() -> stx.CookieManager:
    """A CookieManager is a browser-backed custom component — it must be
    created exactly once per session (re-creating it on every rerun breaks
    its internal state), so it's cached in session_state."""
    if "cookie_manager" not in st.session_state:
        st.session_state["cookie_manager"] = stx.CookieManager(key="furdeco_cookie_manager")
    return st.session_state["cookie_manager"]


def _make_token(password: str, expiry_str: str) -> str:
    """An HMAC of the expiry, keyed by the app password. This lets the
    cookie itself prove it was issued by someone who knew the password,
    without ever storing the password (or anything reversible to it) in
    the cookie."""
    return hmac.new(password.encode(), expiry_str.encode(), hashlib.sha256).hexdigest()


def require_login() -> None:
    """Gate the rest of the app behind a single shared password, since this
    app is publicly reachable and has no per-user accounts. The password
    lives in Streamlit secrets (never in the repo) — locally in
    .streamlit/secrets.toml (gitignored), and on Streamlit Community Cloud
    under the app's Settings -> Secrets.

    A successful login also drops a browser cookie (expiry + HMAC token) so
    the user stays logged in across tab closes / browser restarts for
    COOKIE_EXPIRY_DAYS, instead of having to log in again every visit."""
    if st.session_state.get("authenticated"):
        return

    try:
        correct_password = st.secrets["app_password"]
    except Exception:
        correct_password = None

    if not correct_password:
        st.error(
            "This app has no password configured yet, so it's locked. "
            "Set `app_password` in Secrets (locally: .streamlit/secrets.toml; "
            "on Streamlit Cloud: app Settings -> Secrets) and reload."
        )
        st.stop()

    cookie_manager = get_cookie_manager()

    # Right after an explicit logout, skip trusting the cookie for this one
    # run. The delete() call just told the browser to drop the cookie, but
    # that's an async round trip to the frontend component — a get_all()
    # issued immediately afterward can still read back the old value before
    # the deletion has actually landed, which would silently re-authenticate
    # the user the instant they clicked "Log out". One run without checking
    # the cookie sidesteps that race; by the run after, the deletion has long
    # since landed for real.
    just_logged_out = st.session_state.pop("just_logged_out", False)
    cookie_value = None if just_logged_out else (cookie_manager.get_all() or {}).get(COOKIE_NAME)

    if cookie_value:
        try:
            # "|" rather than ":" — expiry_str is an ISO timestamp and already
            # contains colons (e.g. "2026-09-30T07:49:00"), so splitting on ":"
            # would cut the timestamp itself apart instead of separating it
            # from the token.
            expiry_str, token = cookie_value.split("|", 1)
            expiry = datetime.fromisoformat(expiry_str)
            if expiry > datetime.utcnow() and hmac.compare_digest(token, _make_token(correct_password, expiry_str)):
                st.session_state["authenticated"] = True
                return
        except (ValueError, TypeError):
            pass  # malformed/tampered cookie — fall through to the login form

    st.text_input("Password", type="password", key="login_password")
    submitted = st.button("Log in")

    if submitted:
        entered = st.session_state.get("login_password", "")
        if hmac.compare_digest(entered, correct_password):
            st.session_state["authenticated"] = True
            expiry = datetime.utcnow() + timedelta(days=COOKIE_EXPIRY_DAYS)
            expiry_str = expiry.isoformat()
            token = _make_token(correct_password, expiry_str)
            # Deliberately no st.rerun() here: the cookie component needs this
            # same run to finish rendering in order to actually write the
            # cookie in the browser. Forcing an immediate rerun tears the
            # component down before it gets the chance, so the browser never
            # ends up with the cookie even though session_state looks logged
            # in. Falling through and letting this run finish handles both:
            # session_state is already set, so the app below renders now, and
            # the cookie is left to actually land in the browser for next time.
            cookie_manager.set(
                COOKIE_NAME,
                f"{expiry_str}|{token}",
                expires_at=expiry,
                key="set_furdeco_cookie",
            )
        else:
            st.error("Incorrect password.")

    if not st.session_state.get("authenticated"):
        st.stop()


require_login()

with st.sidebar:
    if st.button("Log out"):
        st.session_state["authenticated"] = False
        st.session_state["just_logged_out"] = True
        get_cookie_manager().delete(COOKIE_NAME, key="delete_furdeco_cookie")
        # A brief pause gives the delete's frontend round-trip a head start
        # before the page reruns (belt-and-braces — require_login()'s
        # just_logged_out check above is what actually guarantees this
        # logout can't be undone by a stale cookie read).
        time.sleep(0.5)
        st.rerun()

uploaded_files = st.file_uploader(
    "Upload invoice PDF(s)",
    type=["pdf"],
    accept_multiple_files=True,
)

if uploaded_files:
    all_rows = []
    errors = []
    price_failure_reasons = {}  # reason -> [order_no, ...], for the diagnostic breakdown below

    with st.spinner("Extracting..."):
        for f in uploaded_files:
            try:
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                    tmp.write(f.getvalue())
                    tmp_path = tmp.name
                try:
                    invoice_no, rows = extract_from_pdf(tmp_path)
                finally:
                    os.unlink(tmp_path)

                if not rows:
                    errors.append(f"{f.name}: no line items found — check the file is a Furdeco invoice.")
                    continue

                for r in rows:
                    amount_str = r["AMOUNT"].replace("£", "").replace(",", "").strip()
                    # Order numbers are meant to be just "CFS" + digits, but
                    # occasionally pick up trailing text after a hyphen (e.g.
                    # "CFS323516-PR9 7JF") from the PDF's own layout — a real
                    # order number never contains a hyphen, so only the part
                    # before it is ever sent to the price API.
                    price_lookup_order_no = (r["ORDER #"] or "").split("-", 1)[0].strip()
                    price, price_fail_reason = fetch_price(price_lookup_order_no)
                    if price_fail_reason:
                        price_failure_reasons.setdefault(price_fail_reason, []).append(r["ORDER #"])
                    all_rows.append(
                        {
                            "DATE": r["DATE"],
                            "RECIPIENT": r["RECIPIENT"],
                            "ORDER #": r["ORDER #"],
                            "POSTCODE": r["POSTCODE"],
                            "DETAILS": r["DETAILS"],
                            "AMOUNT": float(amount_str) if amount_str else 0.0,
                            "Price": price,
                            "Invoice No": invoice_no,
                            "Source File": f.name,
                        }
                    )
            except Exception as e:
                errors.append(f"{f.name}: {e}")

    if errors:
        for e in errors:
            st.warning(e)

    if price_failure_reasons:
        total_failed = sum(len(v) for v in price_failure_reasons.values())
        reason_labels = {
            "empty_order_number": "no Order # was extracted for this row",
            "timeout": "the price API didn't respond in time",
            "connection_error": "couldn't connect to the price API",
            "unparseable_response": "the price API returned something that wasn't valid JSON",
            "api_reported_not_found": "the price API says this order doesn't exist / isn't in its system",
            "malformed_success_response": "the price API said success but the response was missing/invalid data",
        }
        with st.expander(
            f"Price couldn't be looked up for {total_failed} line item(s) — left blank. Click for details.",
            expanded=True,
        ):
            for reason, order_nos in sorted(price_failure_reasons.items(), key=lambda kv: -len(kv[1])):
                label = reason_labels.get(reason, reason) if not reason.startswith("http_") else (
                    f"the price API returned HTTP {reason.removeprefix('http_')}"
                )
                sample = ", ".join(repr(o) for o in order_nos[:8])
                more = f" (+{len(order_nos) - 8} more)" if len(order_nos) > 8 else ""
                st.write(f"**{len(order_nos)}** — {label}: {sample}{more}")

    if all_rows:
        df = pd.DataFrame(all_rows)
        df["Price"] = pd.to_numeric(df["Price"], errors="coerce")  # None -> NaN, so it displays/exports blank rather than the literal text "None"

        total_amount = df["AMOUNT"].sum()
        stat_cols = st.columns(3)
        stat_cols[0].metric("Line items", len(df))
        stat_cols[1].metric("Invoices", df["Invoice No"].nunique())
        stat_cols[2].metric("Total amount", f"£{total_amount:,.2f}")

        st.success(f"Extracted {len(df)} line item(s) from {len(uploaded_files) - len(errors)} PDF(s).")

        # --- On-screen filters (narrow down the preview + the exported Excel) ---
        filterable_cols = ["RECIPIENT", "POSTCODE", "DETAILS", "Invoice No", "Source File"]
        with st.expander("Filter", expanded=False):
            filter_cols = st.columns(len(filterable_cols))
            selections = {}
            for col_widget, col_name in zip(filter_cols, filterable_cols):
                options = sorted(df[col_name].dropna().unique().tolist())
                selections[col_name] = col_widget.multiselect(col_name, options)

        filtered_df = df
        for col_name, chosen in selections.items():
            if chosen:
                filtered_df = filtered_df[filtered_df[col_name].isin(chosen)]

        def _row_style(row):
            # Missing Price is a data-quality flag (the lookup failed) and
            # takes priority over the amount highlight below — it's the more
            # actionable thing to notice, so a row never shows both.
            if pd.isna(row["Price"]):
                return [f"background-color: {LIGHT_BLUE}66"] * len(row)
            if row["AMOUNT"] > HIGHLIGHT_THRESHOLD:
                return [f"background-color: {ROSY_BROWN}55"] * len(row)
            return [""] * len(row)

        display_df = filtered_df[OUTPUT_COLUMNS + ["Source File"]]
        styled_df = (
            display_df.style
            .format({"AMOUNT": "{:.2f}", "Price": "{:.2f}"}, na_rep="")
            .apply(_row_style, axis=1)
        )
        st.dataframe(styled_df, use_container_width=True, hide_index=True)
        st.caption(
            f"Showing {len(filtered_df)} of {len(df)} row(s) — rows with AMOUNT over £{HIGHLIGHT_THRESHOLD} "
            "are highlighted, and rows where Price couldn't be looked up are shown in light blue."
        )

        # Build the Excel file (only the requested columns, in the requested order),
        # with a filter dropdown on every column header so it can be filtered in Excel too.
        export_df = filtered_df[OUTPUT_COLUMNS]
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            export_df.to_excel(writer, index=False, sheet_name="Line Items")
            ws = writer.sheets["Line Items"]
            # Generous widths, and text/number formats and no-wrap alignment set
            # explicitly on every cell — otherwise some viewers (older Excel,
            # LibreOffice with certain locale/default settings) can auto-wrap
            # long values in a narrow column and make rows look "broken" across
            # two visual lines even though the underlying data is a single value.
            widths = [12, 26, 26, 12, 34, 10, 10, 12]
            for col_idx, width in enumerate(widths, start=1):
                ws.column_dimensions[chr(64 + col_idx)].width = width

            last_col_letter = chr(64 + len(OUTPUT_COLUMNS))
            last_row = len(export_df) + 1
            no_wrap = Alignment(wrap_text=False, vertical="center")
            for row in ws.iter_rows(min_row=1, max_row=last_row, max_col=len(OUTPUT_COLUMNS)):
                for cell in row:
                    cell.alignment = no_wrap

            amount_col_letter = chr(64 + OUTPUT_COLUMNS.index("AMOUNT") + 1)
            for cell in ws[amount_col_letter][1:]:
                cell.number_format = "0.00"

            price_col_letter = chr(64 + OUTPUT_COLUMNS.index("Price") + 1)
            for cell in ws[price_col_letter][1:]:
                cell.number_format = "0.00"

            # Highlight rows where AMOUNT is above the threshold, and separately
            # flag rows where Price couldn't be looked up — solid pastel tints
            # (Excel fills don't do translucency reliably across viewers, unlike
            # the on-screen rgba highlight above). A missing Price takes
            # priority, matching the on-screen behaviour above.
            amount_highlight_fill = PatternFill(
                start_color=_blend_with_white(ROSY_BROWN, 0.45),
                end_color=_blend_with_white(ROSY_BROWN, 0.45),
                fill_type="solid",
            )
            missing_price_fill = PatternFill(
                start_color=_blend_with_white(LIGHT_BLUE, 0.6),
                end_color=_blend_with_white(LIGHT_BLUE, 0.6),
                fill_type="solid",
            )
            for row_idx, (amount, price) in enumerate(
                zip(export_df["AMOUNT"], export_df["Price"]), start=2
            ):
                if pd.isna(price):
                    fill = missing_price_fill
                elif amount > HIGHLIGHT_THRESHOLD:
                    fill = amount_highlight_fill
                else:
                    continue
                for cell in ws[row_idx]:
                    cell.fill = fill

            ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
            ws.freeze_panes = "A2"
        buffer.seek(0)

        default_name = f"furdeco_extract_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            "Download Excel",
            data=buffer,
            file_name=default_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    elif not errors:
        st.info("No line items found in the uploaded file(s).")
else:
    st.info("Upload one or more PDFs to get started.")
